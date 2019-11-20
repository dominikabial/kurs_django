from django.core import serializers
import json
from django.core import serializers
from django.utils import timezone
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import Http404, HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, render_to_response
from django.urls import reverse
from django.views import generic
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook

from .models import Book
from .form import CommentForm


# Create your views here.

def book_list(request):
    books = Book.objects.all()
    page = request.GET.get('page')
    paginator = Paginator(books, 10)
    books = paginator.get_page(page)
    format = request.GET.get('format')
    if format == 'json':
        books = serializers.serialize("json", books)
        books = json.loads(books)
        return JsonResponse(books,safe=False)
    if format == 'xml':
        books = serializers.serialize("xml", books)

        return HttpResponse(books, content_type="application/xhtml+xml")
    if format == 'yaml':
        books = serializers.serialize("yaml", books)

        return HttpResponse(books, content_type="application/xhtml+yaml")

    return render(
        request,
        template_name='books/list.html',
        context={'books': books}
    )


class ListView(generic.ListView):
    model = Book
    context_object_name = 'books'
    template_name = 'books/list.html'
    paginate_by = 10

    def get_queryset(self):
        return Book.objects.all()


class DetailView(generic.DetailView, generic.FormView):
    template_name = 'books/detail.html'
    model = Book
    form_class = CommentForm


def book_details(request, book_id):
    form = CommentForm
    try:
        book = Book.objects.get(pk=book_id)
    except Book.DoesNotExist:
        raise Http404("Question does not exist")
    # template = loader.get_template('detail_123.html')
    context = {
        'book': book,
        'form': form
    }
    print(context)
    # return HttpResponse(template.render(context, request))
    return render(
        request,
        template_name='books/detail.html',
        context=context
    )


# POST book/1/comment

@require_http_methods(["POST"])
def book_comment(request, pk):
    form = CommentForm(data=request.POST)
    if form.is_valid():
        comment = form.save(commit=False)
        book = Book.objects.get(pk=pk)
        comment.book = book
        comment.save()
        messages.add_message(request, messages.INFO, 'Comment is add')
        return HttpResponseRedirect(reverse('books:detail', args=(pk,)))
    else:
        messages.add_message(request,messages.WARNING , form.errors)
        return HttpResponseRedirect(reverse('books:detail', args=(pk,)))
#pip install openpyxl
def books_to_excel(request):
    response = HttpResponse(
        content_type='aplication/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition']="attachment; filename=List_{date}.xlsx".format(
        date=timezone.now().strftime("%y-%m-%d")
    )
    workbook = Workbook()
    worksheet = workbook.active
    book_queryset = Book.objects.all()
    paginator = Paginator(book_queryset, 10)
    page = request.GET.get('page')
    book_queryset = paginator.get_page(page)
    worksheet.title = 'Books'
    columns = [
        'ID',
        'Title',
        #'Author'
        'Pages',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for book in book_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            book.id,
            book.title,
            #book.author_text,
            book.pages
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
